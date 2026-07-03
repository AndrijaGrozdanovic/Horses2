import datetime
import pandas as pd
import sys
import subprocess
import time
import re
import openpyxl
from database.DBConnection import DbConnection
from openpyxl.styles import Border, Side
from Selenium.NonRunners import *


def formatFinalReport(path_to_file):
    wb = openpyxl.load_workbook(path_to_file)
    ws = wb.active
    prev_value = ws['C1'].value
    border = Border(bottom=Side(style='thick'))
    for row in range(2, ws.max_row + 1):
        current_value = ws[f'C{row}'].value
        # Check if value has changed
        if current_value != prev_value:
            # Apply thick border to previous row
            ws[f'A{row - 1}'].border = border
            ws[f'B{row - 1}'].border = border
            ws[f'C{row - 1}'].border = border
            ws[f'D{row - 1}'].border = border
            ws[f'E{row - 1}'].border = border
            ws[f'F{row - 1}'].border = border
            # Update previous value
            prev_value = current_value
    wb.save(path_to_file)
    ws = wb.active
    autofit_column_width(ws)
    wb.save(path_to_file)


def autofit_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Excel-style column name
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Padding
        ws.column_dimensions[col_letter].width = adjusted_width


def takeRaceCardDate():
    string = ''
    if int(time.strftime("%H", time.localtime())) > 20:
        Current_Date = datetime.date.today() + datetime.timedelta(days=1)
        string = 'tomorrow'
    else:
        Current_Date = datetime.date.today()
        string = 'today'
    return Current_Date, string


def checkAdditionalCondition(value, column):
    if column == 'SP':
        return f'Odd needs to be {value}'
    elif column == 'favourite':
        return f'Favourite = {value}'
    elif column == 'noOfRunners':
        return column + f" '{value}'"
    else:
        return ''


def checkForParameters(value, column):

    # This block is added for live
    if column == 'SP' or column == 'favourite' or column == 'noOfRunners':
        return '1=1'

    string = column + f" = '{value}'"
    try:
        string = column + '=' + str(float(value))
        return string
    except ValueError:
        pass

    if value == 'ZERO':
        string = column + '=0'
        return string
    if value == 'blanks' and column == 'HG':
        string = "HG=''"
        return string
    if 'Blanks or ' in value:
        string = f"({column} is null or {column} = {re.findall(r'\d+', value)[0]})"

    if value == 'is not null':
        string = f'{column} is not null'

    if value == 'is null':
        string = f'{column} is null'

    if value == '!=1':
        string = f'{column} !=1'

    if column == 'track' or column == 'hType' or column == 'HG' or column == 'trainer' or column == 'Month':
        if ',' in value:
            values = re.split(',', value)
            string = f'{column} in ('
            for val in values:
                string += f"'{val.strip()}',"
            string = string[:-1]
            string += ')'
        else:
            string = column + f" = '{value}'"
    if '>' in str(value) or '<' in str(value) or 'in (' in str(value):
        string = column + f' {value}'
    if 'between' in str(value):
        string = column + f' {value}'
    if column == 'raceName' or column == 'Last_Race_Name':
        string = column + f' {value}'
    if 'not like' in value:
        string = column + f' {value}'
    return string


def executeSystemsInProduction(day, connection):
    file = pd.read_excel('Sistemi_sablon.xlsx').fillna(0)
    for i in file.index:
        where_clause = 'where 1=1 '
        additional = ''
        for column in file.columns:
            if column == 'System' or column == 'Kvota':
                continue
            if file[column][i]:
                where_clause += f'and {checkForParameters(file[column][i], column)} '
                additional += f'  {checkAdditionalCondition(file[column][i], column)}'

        sql = f"select distinct raceDate,Track,horseNameRaw,RaceTime,'{file['System'][i]}' as systemName from Turf_2026_systems "+where_clause
        try:
            result = pd.read_sql(sql=sql, con=connection)
            result.insert(loc=4, column='Condition', value=additional.strip())
            DbConnection.importToTable('system_evidence', connection, result)
        except Exception as e:
            print(f'Error in: {sql}')
            print(e)

    dfr = pd.read_sql('select Track, horseNameRaw, RaceTime, systemName, Condition from system_evidence order by raceTime, horseNameRaw', con=connection)
    path = rf'd:\Konji\Sistemi_2025_evidencija\{day}.xlsx'

    dfr.to_excel(path, index=False)
    formatFinalReport(path)


if __name__ == '__main__':
    days = takeRaceCardDate()
    tomorrow = days[0]
    conObj = DbConnection('mssql')
    conObj.createConnection("AUTOCOMMIT")
    con = conObj.connection
    conObj.executeQuery(BHANonRunners())
    conObj.executeQuery('exec Populate_Overall_MinOR_FTC')
    conObj.executeQuery('exec Populate_OR_Class_L5')
    conObj.executeQuery('exec PRB2_Procedure')
    conObj.executeQuery('exec MedianOR_Prep_procedure')
    conObj.executeQuery('exec CD_Calculations_Procedure')
    conObj.executeQuery('exec Race_Card_enrich')
    conObj.executeQuery('delete from system_evidence')
    executeSystemsInProduction(tomorrow, con)
