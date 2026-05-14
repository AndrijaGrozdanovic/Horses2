import os
import glob
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side

# from LinkCollector.LinkCollector import LinkCollector


# TakeDate = LinkCollector.setRCstartURL().split('/')[4]
path = rf'd:\Konji\Sistemi_2025_evidencija\2026-14-05.xlsx'


def mergeExcel():

    os.chdir(f'd:\\Konji\\2026_Live\\')
    extension = 'csv'
    all_filenames = [i for i in glob.glob('*.{}'.format(extension))]
    combined_csv = pd.concat([pd.read_csv(f) for f in all_filenames])
    sortedExcel = combined_csv.sort_values(by=['RaceTime', 'horseNameRaw'])
    sortedExcel.to_excel(path, index=False)


def autofit_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Excel-style column name
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Padding
        ws.column_dimensions[col_letter].width = adjusted_width


mergeExcel()

wb = openpyxl.load_workbook(path)
ws = wb.active
prev_value = ws['C1'].value
border = Border(bottom=Side(style='thick'))
for row in range(2, ws.max_row + 1):
    current_value = ws[f'C{row}'].value
    # Check if value has changed
    if current_value != prev_value:
        # Apply thick border to previous row
        ws[f'A{row-1}'].border = border
        ws[f'B{row-1}'].border = border
        ws[f'C{row-1}'].border = border
        ws[f'D{row-1}'].border = border
        ws[f'E{row-1}'].border = border
        ws[f'F{row-1}'].border = border
        # Update previous value
        prev_value = current_value
wb.save(path)
ws = wb.active
autofit_column_width(ws)
wb.save(path)